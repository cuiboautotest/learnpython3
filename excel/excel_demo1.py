from openpyxl import Workbook,load_workbook
if __name__ =="__main__":
    print("python openpyxl基本实例")
    #创建excel文档
    wb = Workbook()
    ws =wb.active
    #给默认的工作簿修改名称
    ws.title = "我的默认创建的工作部=簿"
    #对第一行，A-F写入数据
    for col in ("A","B","C","D","E","F"):
        ws["%s1"%col] ="开源优测"

    #对第二行
